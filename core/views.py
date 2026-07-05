from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.shortcuts import get_object_or_404, redirect, render
from django.contrib.auth.decorators import login_required, user_passes_test
from django.contrib.auth.models import User
from django.http import JsonResponse,HttpResponse
from django.views.decorators.http import require_POST
from django.contrib.admin.views.decorators import staff_member_required

from .forms import WorkItemForm
from .models import WorkItem, WorkItemMonthlyValue, WorkItemNote, MonthlyETC, MonthlyETCChange, Project

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.utils import get_column_letter


def is_admin(user):
    return (
        user.is_superuser or
        user.groups.filter(name="admin").exists()
    )

def is_cam(user):
    return (
        is_admin(user) or
        user.groups.filter(name="cam").exists()
    )

def is_read_only(user):
    return user.groups.filter(name="read-only").exists()
    
@staff_member_required
def audit_log(request):
    changes = MonthlyETCChange.objects.select_related(
        "work_item",
        "changed_by"
    ).order_by("-changed_at")[:500]

    return render(request, "core/audit_log.html", {
        "changes": changes
    })

@login_required
@user_passes_test(is_admin)
def user_list(request):
    users = User.objects.all().order_by("username")
    return render(request, "core/user_list.html", {"users": users})

@login_required
def work_item_list(request, project_slug=None):

    project = None

    work_items = (
        WorkItem.objects
        .select_related("project", "assignee")
        .prefetch_related("monthly_values", "monthly_etcs")
    )

    if project_slug:
        project = get_object_or_404(Project, slug=project_slug)
        work_items = work_items.filter(project=project)

    work_items = work_items.order_by("feature_name", "jira_key", "wbs_code")
    work_items = list(work_items)

    db_months_query = WorkItemMonthlyValue.objects

    if project:
        db_months_query = db_months_query.filter(work_item__project=project)

    db_months = list(
        db_months_query
        .values_list("month", flat=True)
        .distinct()
        .order_by("month")
    )

    today = date.today()
    current_month = date(today.year, today.month, 1)

    if db_months:
        start_month = min(db_months[0], current_month)
        
        # Calculate 6 months after the last month in DB
        last_db_month = db_months[-1]
        future_end_month_year = last_db_month.year + (last_db_month.month + 6 - 1) // 12
        future_end_month = date(future_end_month_year, (last_db_month.month + 6 - 1) % 12 + 1, 1)
        
        # Calculate 6 months after current month
        current_future_end_month_year = current_month.year + (current_month.month + 6 - 1) // 12
        current_future_end_month = date(current_future_end_month_year, (current_month.month + 6 - 1) % 12 + 1, 1)
        
        end_month = max(future_end_month, current_future_end_month)
    else:
        start_month = current_month
        # 6 months after current month
        end_month_year = current_month.year + (current_month.month + 6 - 1) // 12
        end_month = date(end_month_year, (current_month.month + 6 - 1) % 12 + 1, 1)

    months = []
    curr = start_month
    while curr <= end_month:
        months.append(curr)
        next_m = curr.month + 1
        next_y = curr.year
        if next_m > 12:
            next_m = 1
            next_y += 1
        curr = date(next_y, next_m, 1)

    grouped_items = defaultdict(list)

    for item in work_items:
        group_key = item.feature_key or "unassigned"

        monthly_lookup = {
            mv.month: mv
            for mv in item.monthly_values.all()
        }

        etc_lookup = {
            etc.month: etc.etc_hours
            for etc in item.monthly_etcs.all()
        }

        item.etc_hours = sum(etc_lookup.values())
        
        item.monthly_display = [
            {
                "month": month,
                "budget": monthly_lookup.get(month).budget_hours if month in monthly_lookup else 0,
                "actual": monthly_lookup.get(month).actual_hours if month in monthly_lookup else 0,
                "etc": etc_lookup.get(month, 0),
            }
            for month in months
        ]

        grouped_items[group_key].append(item)

    feature_groups = []

    for feature_key, items in grouped_items.items():
        first_item = items[0]

        monthly_totals = []
        
        for month in months:
            monthly_totals.append({
                "month": month,
                "budget": sum(row["budget"] for item in items for row in item.monthly_display if row["month"] == month),
                "actual": sum(row["actual"] for item in items for row in item.monthly_display if row["month"] == month),
                "etc": sum(row["etc"] for item in items for row in item.monthly_display if row["month"] == month),
            })

        completed_stories = sum(
            1
            for item in items
            if ((item.jira_status or "").lower() == "done" or (item.jira_status or "").lower() == "void")
        )
        
        total_stories = len(items)
        
        feature_percent_complete = (
            round(100 * completed_stories / total_stories)
            if total_stories > 0
            else 0
        )
        
        feature_baseline_total = sum(item.baseline_budget or 0 for item in items)
        feature_actual_total = sum(month_row.get("actual",0) for month_row in monthly_totals)
        feature_etc_total = sum(item.etc_hours for item in items)
        feature_eac_total = feature_actual_total + feature_etc_total
        feature_variance_total = feature_baseline_total - feature_eac_total

        feature_groups.append({
            "feature_key": feature_key,
            "feature_name": first_item.feature_name or first_item.feature or "Unassigned Feature",
            "feature_url": first_item.feature_url,
            "feature": first_item.feature_name or first_item.feature or "Unassigned Feature",
            "work_item_id": first_item.id,
            "note_text": first_item.note.note_text if hasattr(first_item, "note") else "",
            "items": items,
            "feature_percent_complete": feature_percent_complete,
            "feature_completed_stories": completed_stories,
            "feature_total_stories": total_stories, 
            "budget_total": sum(item.budget_hours for item in items),
            "baseline_total": feature_baseline_total,
            "actual_total": feature_actual_total,
            "etc_total": feature_etc_total,
            "eac_total": feature_eac_total,
            "variance_total": feature_variance_total,
            "monthly_totals": monthly_totals,
        })

    if project:
        project_name = project.name
    else:
        project_name = (
            work_items[0].project.name
            if work_items and work_items[0].project
            else "Project Controls"
        )
    
    projects = Project.objects.all().order_by("name")

    return render(request, "core/work_item_list.html", {
        "feature_groups": feature_groups,
        "months": months,
        "project_name": project_name,
        "projects": projects,
        "current_project": project,
        "is_admin": is_admin(request.user),
    })

@login_required
@user_passes_test(is_cam)
@require_POST
def save_monthly_etc(request):
    work_item_id = request.POST.get("work_item_id")
    month = request.POST.get("month")
    etc_hours = request.POST.get("etc_hours") or 0

    work_item = get_object_or_404(WorkItem, id=work_item_id)
    month_date = datetime.strptime(month, "%Y-%m-%d").date()

    old_etc_hours = None

    existing_entry = MonthlyETC.objects.filter(
        work_item=work_item,
        month=month_date
    ).first()

    if existing_entry:
        old_etc_hours = existing_entry.etc_hours

    entry, created = MonthlyETC.objects.update_or_create(
        work_item=work_item,
        month=month_date,
        defaults={
            "etc_hours": etc_hours,
        }
    )

    if created or str(old_etc_hours) != str(entry.etc_hours):
        MonthlyETCChange.objects.create(
            work_item=work_item,
            month=month_date,
            old_etc_hours=old_etc_hours,
            new_etc_hours=entry.etc_hours,
            changed_by=request.user,
        )

    return JsonResponse({
        "success": True,
        "etc_hours": str(entry.etc_hours),
    })

@login_required
def export_excel(request):

    wb = Workbook()
    ws = wb.active

    ws.append([
        "JIRA Key",
        "Feature Key",
        "Feature",
        "Story",
        "Assignee",
        "Budget",
        "Actuals",
        "ETC",
        "EAC",
        "Variance"
    ])

    work_items = WorkItem.objects.all().order_by("jira_key")\
    
    for item in work_items:
        ws.append([
            item.jira_key,
            item.feature_key,
            item.feature_name,
            item.story,
            (item.assignee.email or item.assignee.username) if item.assignee else "",
            item.budget_hours,
            item.actual_hours,
            item.etc_hours,
            item.eac_hours,
            item.variance_hours
        ])

    # Freeze header row
    ws.freeze_panes = "A2"
    
    # Header style
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    thin_border = Border(
        left=Side(style="thin", color="D9E2F3"),
        right=Side(style="thin", color="D9E2F3"),
        top=Side(style="thin", color="D9E2F3"),
        bottom=Side(style="thin", color="D9E2F3"),
    )
    
    feature_fill = PatternFill("solid", fgColor="D9EAD3")  # feature green
    baseline_fill = PatternFill("solid", fgColor="FCE4D6")  # baseline peach
    jira_fill = PatternFill("solid", fgColor="FFF2CC")      # Jira SP yellow
    actuals_fill = PatternFill("solid", fgColor="D9EAD3")  # actuals green
    etc_fill = PatternFill("solid", fgColor="FCE4D6")      # ETC peach
    variance_fill = PatternFill("solid", fgColor="F4CCCC")  # variance red-ish
    feature_row_fill = PatternFill("solid", fgColor="7F7F7F")
       
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
    
    # Body borders and alignment
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(vertical="center")
    
    # Auto-size columns
    for column_cells in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column_cells[0].column)
    
        for cell in column_cells:
            value = cell.value
            if value is not None:
                max_length = max(max_length, len(str(value)))
    
        ws.column_dimensions[column_letter].width = min(max_length + 2, 35)

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    response["Content-Disposition"] = (
        'attachment; filename="vaporcam_export.xlsx"'
    )
    
    wb.save(response)

    return response

@login_required
def save_note(request):

    if not is_cam(request.user):
        return JsonResponse(
            {"success": False, "error":"Permission denied"},
            status=403
        )
        
    if request.method != "POST":
        return JsonResponse({"success": False}, status=405)

    work_item_id = request.POST.get("work_item_id")
    note_text = request.POST.get("note_text", "")

    work_item = WorkItem.objects.get(id=work_item_id)

    note, created = WorkItemNote.objects.get_or_create(
        work_item=work_item
    )

    note.note_text = note_text
    note.save()

    return JsonResponse({"success": True})
    
@login_required
@user_passes_test(is_admin)
def update_baseline(request):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "POST required"}, status=405)

    work_item_id = request.POST.get("work_item_id")
    baseline_budget = request.POST.get("baseline_budget", "0")

    if not work_item_id:
        return JsonResponse({"success": False, "error": "Missing work item ID"}, status=400)

    try:
        baseline_value = Decimal(baseline_budget)
    except InvalidOperation:
        return JsonResponse({"success": False, "error": "Invalid baseline value"}, status=400)

    try:
        work_item = WorkItem.objects.get(id=work_item_id)
    except WorkItem.DoesNotExist:
        return JsonResponse({"success": False, "error": "Work item not found"}, status=404)

    work_item.baseline_budget = baseline_value
    work_item.save()

    return JsonResponse({
        "success": True,
        "baseline_budget": str(work_item.baseline_budget),
    })
