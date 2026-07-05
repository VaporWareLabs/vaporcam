import openpyxl

from collections import defaultdict
from datetime import date, datetime
from decimal import Decimal, InvalidOperation

from django.apps import apps
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils.dateparse import parse_date

from core.models import WorkItem, WorkItemMonthlyValue


CHARGE_CODE_HEADERS = (
    "PROJECT ID",
    "CHARGE CODE",
    "CHARGECODE",
    "WBS",
)

HOURS_HEADERS = (
    "ENTERED HOURS",
    "TOTAL(ENTERED HOURS)",
    "HOURS",
)

MONTH_HEADERS = (
    "HOURS DATE",
    "HOUR DATE",
    "DATE",
    "MONTH",
    "PERIOD",
)

DEFAULT_START_MONTH = "2026-02"

def normalize(value):
    return str(value or "").strip().upper()


def to_decimal(value):
    if value in (None, ""):
        return Decimal("0")

    try:
        return Decimal(str(value)).quantize(Decimal("0.01"))
    except (InvalidOperation, TypeError, ValueError):
        return Decimal("0")


def first_matching_header(headers, possible_names):
    for name in possible_names:
        col = headers.get(normalize(name))
        if col:
            return col, name

    return None, None


def month_start(value):
    """
    Convert an Excel date/month value into the first day of that month.

    Supports real Excel dates, strings like 2026-06-15, 2026-06, 06/2026,
    and common month labels like Jun-2026.
    """
    if value in (None, ""):
        return None

    if isinstance(value, datetime):
        return date(value.year, value.month, 1)

    if isinstance(value, date):
        return date(value.year, value.month, 1)

    text = str(value).strip()
    if not text:
        return None

    # YYYY-MM is a common month-only format; parse_date needs a day.
    if len(text) == 7 and text[4] == "-":
        parsed = parse_date(f"{text}-01")
        if parsed:
            return date(parsed.year, parsed.month, 1)

    parsed = parse_date(text)
    if parsed:
        return date(parsed.year, parsed.month, 1)

    # Last resort for strings like 06/2026, 6/2026, Jun-2026, June 2026.
    for fmt in ("%m/%Y", "%m-%Y", "%b-%Y", "%B %Y", "%b %Y"):
        try:
            parsed_dt = datetime.strptime(text, fmt)
            return date(parsed_dt.year, parsed_dt.month, 1)
        except ValueError:
            pass

    return None


def get_monthly_actual_model():
    """
    Return core.MonthlyActual if it exists.

    This lets the command fail gracefully if the model/migration has not been
    added yet, instead of crashing at import time.
    """
    try:
        return apps.get_model("core", "MonthlyActual")
    except LookupError:
        return None


class Command(BaseCommand):
    help = "Import monthly actual hours from an Excel actuals report"

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", type=str)
        parser.add_argument(
            "--dry-run",
            action="store_true",
            help="Preview import results without updating the database",
        )
        parser.add_argument(
            "--month",
            type=str,
            help="Month to apply all rows to when the spreadsheet has no month/date column. Format: YYYY-MM",
        )
        parser.add_argument(
            "--sheet",
            type=str,
            help="Worksheet name. Defaults to the active worksheet.",
        )

    def handle(self, *args, **options):
        xlsx_path = options["xlsx_path"]
        dry_run = options["dry_run"]
        default_month = month_start(options.get("month")) if options.get("month") else None

        start_month = datetime.strptime(
            DEFAULT_START_MONTH,
            "%Y-%m"
        ).date().replace(day=1)
       
        if options.get("month") and not default_month:
            raise CommandError("Invalid --month value. Use YYYY-MM, for example: --month 2026-06")

        wb = openpyxl.load_workbook(xlsx_path, data_only=True)

        if options.get("sheet"):
            if options["sheet"] not in wb.sheetnames:
                raise CommandError(f"Worksheet not found: {options['sheet']}")
            ws = wb[options["sheet"]]
        else:
            ws = wb.active

        headers = {
            normalize(cell.value): idx
            for idx, cell in enumerate(ws[1], start=1)
            if normalize(cell.value)
        }

        charge_col, charge_header = first_matching_header(headers, CHARGE_CODE_HEADERS)
        hours_col, hours_header = first_matching_header(headers, HOURS_HEADERS)
        month_col, month_header = first_matching_header(headers, MONTH_HEADERS)

        missing = []
        if not charge_col:
            missing.append("charge code column, such as PROJECT ID")
        if not hours_col:
            missing.append("hours column, such as ENTERED HOURS")
        if not month_col and not default_month:
            missing.append("month/date column, or provide --month YYYY-MM")

        if missing:
            self.stderr.write(self.style.ERROR("Could not find required input data."))
            for item in missing:
                self.stderr.write(f"  Missing: {item}")

            self.stdout.write("")
            self.stdout.write("Headers found:")
            for header in headers:
                self.stdout.write(f"  {header}")

            return

        self.stdout.write(f"Using charge code column: {charge_header}")
        self.stdout.write(f"Using hours column: {hours_header}")
        if month_col:
            self.stdout.write(f"Using month/date column: {month_header}")
        else:
            self.stdout.write(f"Using default month: {default_month:%Y-%m}")

        actuals_by_charge_code_month = defaultdict(Decimal)
        invalid_month_rows = []

        rows_processed = 0
        rows_with_hours = 0

        for row_num in range(2, ws.max_row + 1):
            rows_processed += 1

            charge_code = normalize(ws.cell(row=row_num, column=charge_col).value)
            hours = to_decimal(ws.cell(row=row_num, column=hours_col).value)

            if not charge_code or hours == 0:
                continue

            if month_col:
                actual_month = month_start(ws.cell(row=row_num, column=month_col).value)
            else:
                actual_month = default_month

            if not actual_month:
                invalid_month_rows.append(row_num)
                continue

            #-------------------------------------------
            # Ingnore months before cutoff
            #-------------------------------------------
            if start_month and actual_month < start_month:
                continue

            rows_with_hours += 1
            actuals_by_charge_code_month[(charge_code, actual_month)] += hours

        total_file_hours = sum(actuals_by_charge_code_month.values())
        charge_codes_in_file = {code for code, _month in actuals_by_charge_code_month.keys()}
        months_in_file = sorted({_month for _code, _month in actuals_by_charge_code_month.keys()})

        self.stdout.write("")
        self.stdout.write(
            self.style.SUCCESS(
                f"Found actuals for {len(charge_codes_in_file)} charge codes across {len(months_in_file)} month(s)."
            )
        )

        if invalid_month_rows:
            self.stdout.write(
                self.style.WARNING(
                    f"Skipped {len(invalid_month_rows)} row(s) because the month/date could not be parsed."
                )
            )

        feature_charge_codes = defaultdict(set)
        matched_items = 0

        work_items = (
            WorkItem.objects
            .exclude(charge_code__isnull=True)
            .exclude(charge_code="")
        )

        for item in work_items:
            charge_code = normalize(item.charge_code)

            if charge_code not in charge_codes_in_file:
                continue

            feature_key = normalize(item.feature_key)

            if not feature_key:
                self.stdout.write(
                    self.style.WARNING(
                        f"Skipping {item.jira_key}: matched charge code but no feature_key"
                    )
                )
                continue

            feature_charge_codes[feature_key].add(charge_code)
            matched_items += 1

        feature_actuals_by_month = defaultdict(Decimal)

        for feature_key, charge_codes in feature_charge_codes.items():
            for charge_code in charge_codes:
                for (actual_charge_code, actual_month), hours in actuals_by_charge_code_month.items():
                    if actual_charge_code == charge_code:
                        feature_actuals_by_month[(feature_key, actual_month)] += hours

        matched_charge_code_count = len(
            {
                charge_code
                for charge_codes in feature_charge_codes.values()
                for charge_code in charge_codes
            }
        )

        total_matched_hours = sum(feature_actuals_by_month.values())
        total_unmatched_hours = total_file_hours - total_matched_hours
        percent_matched = (
            (total_matched_hours / total_file_hours) * 100
            if total_file_hours
            else 0
        )

        self.stdout.write("")
        self.stdout.write("Actuals Import Summary")
        self.stdout.write("----------------------")
        self.stdout.write(f"Rows processed: {rows_processed}")
        self.stdout.write(f"Rows with hours: {rows_with_hours}")
        self.stdout.write(f"Charge codes in file: {len(charge_codes_in_file)}")
        self.stdout.write(f"Charge codes matched: {matched_charge_code_count}")
        self.stdout.write(f"Months in file: {', '.join(month.strftime('%Y-%m') for month in months_in_file)}")
        self.stdout.write(f"Total hours in file: {total_file_hours}")
        self.stdout.write(f"Total matched hours: {total_matched_hours}")
        self.stdout.write(f"Unmatched hours: {total_unmatched_hours}")
        self.stdout.write(f"Percent matched: {percent_matched:.1f}%")
        self.stdout.write("")

        unmatched_charge_codes = sorted(charge_codes_in_file - {
            charge_code
            for charge_codes in feature_charge_codes.values()
            for charge_code in charge_codes
        })

        if unmatched_charge_codes:
            self.stdout.write("Unmatched charge codes:")
            for charge_code in unmatched_charge_codes:
                hours = sum(
                    actuals_by_charge_code_month[(charge_code, actual_month)]
                    for actual_month in months_in_file
                    if (charge_code, actual_month) in actuals_by_charge_code_month
                )
                self.stdout.write(f"  {charge_code}: {hours} hrs")
            self.stdout.write("")

        if dry_run:
            self.stdout.write(self.style.WARNING("DRY RUN: No database changes were made."))
        else:
            WorkItemMonthlyValue.objects.update(actual_hours=0)
        
        updated_monthly_values = 0
        
        for (feature_key, actual_month), actual_hours in feature_actuals_by_month.items():
            feature_items = WorkItem.objects.filter(
                feature_key=feature_key
            ).order_by("jira_key")
        
            feature_item = feature_items.first()
        
            if not feature_item:
                self.stdout.write(
                    self.style.WARNING(
                        f"No WorkItem found for feature_key {feature_key}"
                    )
                )
                continue
        
            if not dry_run:
                monthly_value, _created = WorkItemMonthlyValue.objects.get_or_create(
                    work_item=feature_item,
                    month=actual_month,
                    defaults={
                        "budget_hours": Decimal("0"),
                        "actual_hours": Decimal("0"),
                    },
                )
        
                monthly_value.actual_hours = actual_hours
                monthly_value.save(update_fields=["actual_hours"])
        
            updated_monthly_values += 1
        
            self.stdout.write(
                f"{feature_key} {actual_month:%Y-%m}: {actual_hours} hrs applied to {feature_item.jira_key}"
            )

